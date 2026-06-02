"""
报告生成工具模块
支持PDF和Excel格式报告
"""

import os
import logging
from datetime import datetime
from typing import Dict, List
from django.conf import settings

logger = logging.getLogger(__name__)


class ReportGenerator:
    """报告生成器"""

    def __init__(self):
        self.output_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_pdf(self, data: Dict, report_id: int) -> str:
        """生成PDF报告"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm, mm
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.colors import HexColor
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            logger.error("reportlab未安装，无法生成PDF报告")
            raise

        filename = f"report_{report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.output_dir, filename)

        # 注册中文字体（如果可用）
        try:
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
            font_name = 'STSong-Light'
        except Exception:
            font_name = 'Helvetica'

        # 创建PDF
        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        # 样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontName=font_name,
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10
        )
        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            spaceBefore=5,
            spaceAfter=5
        )

        # 构建内容
        story = []

        # 标题
        story.append(Paragraph(f"光伏行业舆情与股价分析报告", title_style))
        story.append(Spacer(1, 20))

        # 股票信息
        stock_info = data.get('stock', {})
        story.append(Paragraph(f"股票信息", heading_style))
        story.append(Paragraph(f"股票代码：{stock_info.get('code', 'N/A')}", body_style))
        story.append(Paragraph(f"股票名称：{stock_info.get('name', 'N/A')}", body_style))
        story.append(Paragraph(f"所属行业：{stock_info.get('industry', 'N/A')}", body_style))
        story.append(Paragraph(f"分析周期：{data.get('period', 'N/A')}", body_style))
        story.append(Spacer(1, 20))

        # 分析结果
        analysis = data.get('analysis', [])
        if analysis:
            story.append(Paragraph("分析结果", heading_style))
            for a in analysis:
                story.append(Paragraph(f"• {a.get('type', '')}: {a.get('summary', '')}", body_style))
            story.append(Spacer(1, 20))

        # 情感趋势
        sentiment_trend = data.get('sentiment_trend', [])
        if sentiment_trend:
            story.append(Paragraph("情感趋势", heading_style))

            # 创建表格
            table_data = [['日期', '情感得分', '新闻数量']]
            for s in sentiment_trend[:10]:  # 只显示前10条
                table_data.append([
                    s.get('date', ''),
                    s.get('score', ''),
                    str(s.get('count', 0))
                ])

            table = Table(table_data, colWidths=[4*cm, 4*cm, 4*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, HexColor('#D9D9D9')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FFFFFF'), HexColor('#F2F2F2')]),
            ]))
            story.append(table)
            story.append(Spacer(1, 20))

        # 新闻列表
        news = data.get('news', [])
        if news:
            story.append(Paragraph("相关新闻", heading_style))
            for n in news[:10]:
                sentiment = n.get('sentiment', 'neutral')
                sentiment_map = {'positive': '正面', 'negative': '负面', 'neutral': '中性'}
                story.append(Paragraph(
                    f"• [{n.get('date', '')}] [{sentiment_map.get(sentiment, '中性')}] {n.get('title', '')}",
                    body_style
                ))
            story.append(Spacer(1, 20))

        # 免责声明
        story.append(Spacer(1, 40))
        disclaimer_style = ParagraphStyle(
            'Disclaimer',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8,
            textColor=HexColor('#999999')
        )
        story.append(Paragraph(
            "免责声明：本报告仅供参考，不构成投资建议。投资有风险，入市需谨慎。",
            disclaimer_style
        ))
        story.append(Paragraph(
            f"报告生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            disclaimer_style
        ))

        # 生成PDF
        doc.build(story)

        return f"reports/{filename}"

    def generate_excel(self, data: Dict, report_id: int) -> str:
        """生成Excel报告"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            logger.error("openpyxl未安装，无法生成Excel报告")
            raise

        filename = f"report_{report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()

        # 样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: 股票信息
        ws1 = wb.active
        ws1.title = "股票信息"
        ws1.append(["股票代码", "股票名称", "所属行业", "分析周期"])
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        stock_info = data.get('stock', {})
        ws1.append([
            stock_info.get('code', ''),
            stock_info.get('name', ''),
            stock_info.get('industry', ''),
            data.get('period', '')
        ])

        # Sheet 2: 行情数据
        ws2 = wb.create_sheet("行情数据")
        headers = ['日期', '开盘价', '收盘价', '最高价', '最低价', '成交量', '涨跌幅(%)']
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for q in data.get('quotes', []):
            ws2.append([
                q.get('date', ''),
                q.get('open', ''),
                q.get('close', ''),
                q.get('high', ''),
                q.get('low', ''),
                q.get('volume', 0),
                q.get('change', '')
            ])

        # Sheet 3: 情感趋势
        ws3 = wb.create_sheet("情感趋势")
        ws3.append(['日期', '情感得分', '新闻数量'])
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for s in data.get('sentiment_trend', []):
            ws3.append([
                s.get('date', ''),
                s.get('score', ''),
                s.get('count', 0)
            ])

        # Sheet 4: 新闻列表
        ws4 = wb.create_sheet("新闻列表")
        ws4.append(['日期', '标题', '来源', '情感', '得分'])
        for cell in ws4[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        sentiment_map = {'positive': '正面', 'negative': '负面', 'neutral': '中性'}
        for n in data.get('news', []):
            ws4.append([
                n.get('date', ''),
                n.get('title', ''),
                n.get('source', ''),
                sentiment_map.get(n.get('sentiment', 'neutral'), '中性'),
                n.get('score', '')
            ])

        # 调整列宽
        for ws in [ws1, ws2, ws3, ws4]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)

        return f"reports/{filename}"
